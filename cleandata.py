import streamlit as st
import pandas as pd
import openpyxl
import json
from openpyxl.utils import get_column_letter

# Required columns to keep
REQUIRED_COLS = [
    "Academic Level",
    "Primary Program of Study",
    "Current Primary College / School",
    "Cumulative GPA",
    "Gender",
    "FTIC Cohort",
    "Cumulative Credits: Earned"
]

def extract_filters(file_path):
    """Extract Excel AutoFilters from the file"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    if not ws.auto_filter or not ws.auto_filter.ref:
        return {}
    
    filters_dict = {}
    
    for flt in ws.auto_filter.filterColumn:
        col_letter = get_column_letter(flt.colId + 1)
        allowed = []
        
        for f in flt.filters:
            if hasattr(f, "val"):
                allowed.append(f.val)
        
        filters_dict[col_letter] = allowed
    
    return filters_dict

st.title("📊 Excel Data Cleaner")
st.write("Upload your Excel file to clean it.")
st.info("The first 20 rows will be skipped. Row 21 becomes the header.")

uploaded = st.file_uploader("Upload Excel File (.xlsx)", type=["xlsx"])

if uploaded:
    st.success("✅ File uploaded successfully!")

    # Save temp file
    temp_path = "temp.xlsx"
    with open(temp_path, "wb") as f:
        f.write(uploaded.getbuffer())

    try:
        # Extract filters
        filters = extract_filters(temp_path)
        
        if filters:
            st.subheader("📘 Extracted Excel Filters")
            st.json(filters)
            # Save filters in session state
            st.session_state.extracted_filters = filters
            # Save to JSON file
            with open("excel_filters.json", "w") as f:
                json.dump(filters, f, indent=4)
            st.success("✅ Filters saved to excel_filters.json")
        else:
            st.info("No filters found in the Excel file.")
        
        # Read Excel: skip first 20 rows, row 21 is header
        df = pd.read_excel(temp_path, skiprows=20)
        
        # Clean column names - strip whitespace
        df.columns = df.columns.str.strip()
        
        # Check for required columns
        missing = [col for col in REQUIRED_COLS if col not in df.columns]
        
        if missing:
            st.error(f"Missing columns: {', '.join(missing)}")
            st.write(f"Available columns: {list(df.columns)}")
        else:
            # Keep only required columns
            df = df[REQUIRED_COLS]
            
            # Remove completely empty rows (all columns are null)
            df = df.dropna(how="all")
            
            # Reset index to start from 1
            df = df.reset_index(drop=True)
            df.index = df.index + 1
            df.index.name = "Student #"
            
            st.subheader("Cleaned Data")
            st.dataframe(df)
            
            st.success(f"Total students: **{len(df)}**")
            
            # Download button
            xlsx_path = "cleaned_output.xlsx"
            df.to_excel(xlsx_path, index=True)
            
            with open(xlsx_path, "rb") as f:
                st.download_button(
                    label="Download Cleaned XLSX",
                    data=f,
                    file_name="cleaned_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"❌ Error: {str(e)}")