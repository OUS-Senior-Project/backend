import json
import sys

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# REQUIRED COLUMNS TO EXTRACT (user defined)
# ------------------------------------------------------------
TARGET_COLS = [
    "Academic Level",
    "Overall Registration Status",
    "Current Academic Record Status",
    "Current Primary Academic Unit",
    "Latest Class Standing",
    "Current Primary Program of Study",
    "First Standard Academic Period for Student",
    "Load Status",
    "Program of Study (All Periods)",
    "Student Cohorts",
    "Student",
    "Student ID",
    "Veterans Bill /Benefit",
    "Currently Enrolled",
    "Is International Student",
    "Starting Academic Period Start Date",
    "Ending Academic Period End Date",
]


# ------------------------------------------------------------
# Extract Excel AutoFilters using openpyxl
# ------------------------------------------------------------
def extract_filters(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    if not ws.auto_filter or not ws.auto_filter.ref:
        return {}

    filters_dict = {}

    for flt in ws.auto_filter.filterColumn:
        col_letter = get_column_letter(flt.colId + 1)
        allowed = []

        for f in flt.filters:
            if hasattr(f, "val"):
                allowed.append(f.val)

        filters_dict[col_letter] = allowed

    return filters_dict


# ------------------------------------------------------------
# Smart column matcher (contains instead of startswith)
# ------------------------------------------------------------
def match_target_columns(df):
    matched = {}
    missing = []

    for col_prefix in TARGET_COLS:
        found = None
        prefix_lower = col_prefix.lower()

        # Try to match if prefix appears anywhere inside column name
        for col in df.columns:
            if prefix_lower in str(col).lower():
                found = col
                break

        if found:
            matched[col_prefix] = found
        else:
            missing.append(col_prefix)

    return matched, missing


# ------------------------------------------------------------
# Clean Data Function
# ------------------------------------------------------------
def clean_data(df):
    # --- 1. Auto detect header row ---
    header_row = None
    for i in range(0, 30):
        # A valid header row usually has many non-empty cells
        if df.iloc[i].count() >= 10:
            header_row = i
            break

    if header_row is None:
        raise Exception("Could not detect header row automatically.")

    print(f"Detected header row at index {header_row}")

    # --- 2. Apply header ---
    df = df.iloc[header_row:].reset_index(drop=True)
    df.columns = df.iloc[0]
    df = df[1:].reset_index(drop=True)

    # --- DEBUG ---
    print("\n--- Columns detected after header extraction ---")
    for c in df.columns:
        print(f"- {c}")
    print("----------------------------------------------\n")

    # --- 3. Match required columns ---
    matched_cols, missing = match_target_columns(df)

    if missing:
        raise Exception("Missing columns:\n" + "\n".join("- " + m for m in missing))

    # --- 4. Reduce to needed columns ---
    df = df[[matched_cols[p] for p in TARGET_COLS]]

    # --- 5. Standardize names ---
    df = df.rename(columns={matched_cols[p]: p for p in TARGET_COLS})

    # --- 6. Remove blank rows ---
    df = df.dropna(how="all")

    return df


# ------------------------------------------------------------
# MAIN CLI ENTRY
# ------------------------------------------------------------
def main():
    if len(sys.argv) < 2:
        print("Usage: python clicleandata.py <excel_file.xlsx>")
        sys.exit(1)

    excel_path = sys.argv[1]
    print(f"Loading: {excel_path}")

    # ---------------------------
    # Extract filters
    # ---------------------------
    print("Extracting filters...")
    filters = extract_filters(excel_path)

    with open("excel_filters.json", "w") as f:
        json.dump(filters, f, indent=4)

    print("Saved excel_filters.json")

    # ---------------------------
    # Load Workbook
    # ---------------------------
    print("Cleaning workbook data...")
    raw_df = pd.read_excel(excel_path, engine="openpyxl")

    cleaned_df = clean_data(raw_df)

    # ---------------------------
    # Save cleaned dataset
    # ---------------------------
    cleaned_df.to_csv("cleaned_output.csv", index=False)
    print("Saved cleaned_output.csv")

    print("\nProcessing complete!")


if __name__ == "__main__":
    main()
